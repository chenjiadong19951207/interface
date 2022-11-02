#coding=utf8
import openpyxl
from openpyxl import workbook
import sys
from collections.abc import Iterable
import os
base_path = os.getcwd()
sys.path.append(base_path)

class HandExcel:
    def load_excel(self):
        '''加载excel'''
        open_excel = openpyxl.load_workbook(base_path+'/Case/imooc.xlsx')
        return open_excel

    def get_sheet_data(self,index=None):
        '''加载所有sheet的内容'''
        sheet_name = self.load_excel().sheetnames
        if index ==None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,cols):
        '''获取某一个单元格内容'''
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data
